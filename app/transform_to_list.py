import os 
import json
import openpyxl
from bs4 import BeautifulSoup


class Transform:
    def __init__(self, filename, route = None):
        self.file = filename
        self.route = '.' + route
        self.extension = self.file.split('.')[-1]
        self._path = self._absolute_path()
        self.list_values = self._proceed()
        
    def list_from_csv(self):
        rows_list = []
        path = self._absolute_path()
        with open(self._path, 'r') as file:
            data = file.readlines()

            for line in data:
                line = line.replace('\n', '')
                rows_list.append(line.split(','))

        return rows_list
    
    def list_from_xlsx(self):
        file = openpyxl.load_workbook(self._path)
        sheet = file.active
        rows_list = []

        for row in sheet.iter_rows(max_row = sheet.max_row):
            value_list = []
            
            for cell in row:
                value_list.append(cell.value)
            rows_list.append(value_list)
        
        return rows_list
    
    def list_from_json(self):
        rows_list = []
        with open(self._path, 'r') as file:
            data = json.load(file)
            rows_list.append(list(data[1].keys()))
            for line in data:
                rows_list.append(list(line.values()))
            
        return rows_list

    def list_from_html(self):
        rows_list = []
        with open(self._path, 'r') as file:
            data = file.read()
            
            soup = BeautifulSoup(data, 'lxml')
            rows = soup.find('table').find_all('tr')

            for row in rows:
                value_list = []
                lines = row.find_all(['td', 'th'])

                for line in lines:
                    value_list.append(line.getText())

                rows_list.append(value_list)

            return rows_list



    def _proceed(self):
        '''Select the procedure to each of the valid extensions'''
        if self.extension == 'csv':
            return self.list_from_csv()
        elif self.extension == 'xlsx':
            return  self.list_from_xlsx()
        elif self.extension == 'json':
            return  self.list_from_json()
        elif self.extension == 'html':
            return self.list_from_html()
        else:
            return None

        
                
    def _absolute_path(self):
        '''Define the absolute path to the file'''
        if self.route is not None:
            #Asolute path to the current script
            my_path = os.path.abspath(os.path.dirname(__file__))
            path_relative_file = os.path.join(self.route ,  self.file)
            #Absolute path to the data file
            path = os.path.join(my_path, path_relative_file)
            return path
        else:
            return None


    



        
